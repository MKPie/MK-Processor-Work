#!/usr/bin/env python3
import sys
import os
import pandas as pd
import gspread
import re
import math
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, QPushButton,
    QProgressBar, QScrollArea, QFrame, QMessageBox, QComboBox
)
from PyQt5.QtCore import Qt, QTimer, pyqtSignal, QObject
from PyQt5.QtGui import QFont
from oauth2client.service_account import ServiceAccountCredentials
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from fake_useragent import UserAgent
import threading
import time
import traceback
import openpyxl
from openpyxl.styles import Alignment

# Add these new imports for our extensions
from config_manager import ConfigManager
from settings_dialog import SettingsDialog
from plugin_manager import PluginManager
from webscraper_wrapper import create_webscraper_wrapper
from excel_formatter import enhance_save_results
from plugin_manager_dialog import PluginManagerDialog

# Simple class for better error handling
class AppError(Exception):
    pass

# Define a signal class for thread-safe GUI updates
class WorkerSignals(QObject):
    update_progress = pyqtSignal(int, int)
    update_status = pyqtSignal(str)
    finished = pyqtSignal()
    error = pyqtSignal(str)

class SheetRow(QFrame):
    def __init__(self, index, parent):
        super().__init__(parent)
        self.index = index
        self.parent = parent
        self.running = False
        self.completed = False
        self.output_df = None
        self.output_path = None
        self.selected_file = None
        self.worker_thread = None
        self.signals = WorkerSignals()
        
        # Set up UI
        self.setup_ui()
        
        # Connect signals for thread-safe updates
        self.signals.update_progress.connect(self.on_update_progress)
        self.signals.update_status.connect(self.on_update_status)
        self.signals.finished.connect(self.on_processing_finished)
        self.signals.error.connect(self.on_processing_error)
        
        # Load files in dropdown
        QTimer.singleShot(500, self.load_files)
    
    def setup_ui(self):
        # Basic styling
        self.setFrameShape(QFrame.StyledPanel)
        self.setStyleSheet("""
            QFrame {
                background-color: white;
                border-radius: 2px;
                padding: 8px;
                border: 1px solid #cccccc;
            }
            QProgressBar {
                border: 1px solid #cccccc;
                border-radius: 2px;
                text-align: center;
                height: 14px;
                font-size: 9px;
                margin-top: 0px;
            }
            QProgressBar::chunk {
                background-color: #4285f4;
                border-radius: 1px;
            }
            QComboBox {
                border: 1px solid #aaaaaa;
                border-radius: 2px;
                padding: 3px;
                background-color: white;
                min-height: 24px;
                max-height: 24px;
                font-size: 11px;
            }
            QLineEdit {
                border: 1px solid #aaaaaa;
                border-radius: 2px;
                padding: 3px;
                background-color: white;
                min-height: 24px;
                max-height: 24px;
                font-size: 11px;
            }
            QPushButton {
                border-radius: 2px;
                padding: 3px;
                min-height: 24px;
                max-height: 24px;
                font-weight: bold;
                font-size: 11px;
            }
            QPushButton#startBtn {
                background-color: #4285f4;
                color: white;
                border: none;
            }
            QPushButton#startBtn:hover {
                background-color: #3367d6;
            }
            QPushButton#startBtn:disabled {
                background-color: #a5c2f5;
            }
            QPushButton#stopBtn {
                background-color: #f0f0f0;
                color: #333333;
                border: 1px solid #cccccc;
            }
            QPushButton#stopBtn:hover {
                background-color: #e0e0e0;
            }
            QPushButton#stopBtn:disabled {
                color: #aaaaaa;
                border: 1px solid #e0e0e0;
            }
            QPushButton#refreshBtn {
                background-color: #f0f0f0;
                color: #333333;
                border: 1px solid #aaaaaa;
                font-size: 13px;
                min-width: 24px;
                max-width: 24px;
            }
            QPushButton#refreshBtn:hover {
                background-color: #e0e0e0;
            }
            QLabel {
                color: #333333;
                font-size: 11px;
            }
        """)
        
        # Set object name for debugging
        self.setObjectName(f"SheetRow_{self.index}")
        
        # Main layout
        layout = QVBoxLayout(self)
        layout.setSpacing(6)  # Further reduced spacing between sections
        layout.setContentsMargins(10, 8, 10, 8)  # Smaller margins
        
        # Top row: file selection and buttons
        top_row = QHBoxLayout()
        top_row.setSpacing(4)  # Tighter spacing between controls
        
        # File dropdown
        self.file_dropdown = QComboBox(self)
        self.file_dropdown.setFixedHeight(24)  # Smaller height
        self.file_dropdown.addItem("")
        self.file_dropdown.setPlaceholderText("Select File")
        
        # Connect selection change event
        self.file_dropdown.currentIndexChanged.connect(self.file_selected)
        
        # Refresh button
        self.refresh_btn = QPushButton("↻", self)
        self.refresh_btn.setObjectName("refreshBtn")
        self.refresh_btn.setFixedSize(24, 24)  # Smaller square button
        self.refresh_btn.clicked.connect(self.load_files)
        
        # Prefix input
        self.prefix_input = QLineEdit(self)
        self.prefix_input.setPlaceholderText("Prefix")
        self.prefix_input.setFixedSize(100, 24)  # Smaller size
        
        # Start button
        self.start_btn = QPushButton("Start", self)
        self.start_btn.setObjectName("startBtn")
        self.start_btn.setFixedSize(80, 24)  # Smaller size
        self.start_btn.clicked.connect(self.start_processing)
        
        # Stop button
        self.stop_btn = QPushButton("Stop", self)
        self.stop_btn.setObjectName("stopBtn")
        self.stop_btn.setFixedSize(80, 24)  # Smaller size
        self.stop_btn.clicked.connect(self.stop_processing)
        self.stop_btn.setEnabled(False)
        
        # Add widgets to top row with exact spacing
        top_row.addWidget(self.file_dropdown, 1)  # Stretch factor 1
        top_row.addWidget(self.refresh_btn)
        top_row.addSpacing(4)  # Tighter spacing
        top_row.addWidget(self.prefix_input)
        top_row.addSpacing(4)  # Tighter spacing
        top_row.addWidget(self.start_btn)
        top_row.addSpacing(4)  # Tighter spacing
        top_row.addWidget(self.stop_btn)
        
        # Middle row: Status information
        status_row = QHBoxLayout()
        status_row.setContentsMargins(0, 3, 0, 3)  # Minimal vertical padding
        
        self.status_label_prefix = QLabel("Status:", self)
        self.status_label_prefix.setFont(QFont("Arial", 8, QFont.Bold))
        
        self.status_label = QLabel("Ready", self)
        
        status_row.addWidget(self.status_label_prefix)
        status_row.addWidget(self.status_label)
        status_row.addStretch(1)
        
        # Bottom row: Progress bar
        progress_row = QHBoxLayout()
        
        self.progress_bar = QProgressBar(self)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setFixedHeight(14)  # Even smaller height
        
        progress_row.addWidget(self.progress_bar)
        
        # Add layouts to main layout with reduced spacing
        layout.addLayout(top_row)
        layout.addLayout(status_row)
        layout.addLayout(progress_row)
    
    # Signal handler methods for thread-safe UI updates
    def on_update_progress(self, current, total):
        if total <= 0:
            percent = 0
        else:
            percent = int((current / total) * 100)
        
        self.progress_bar.setValue(percent)
        
        # Also update parent window info
        selected_file = self.get_selected_file()
        if selected_file:
            self.parent.update_status(f"Processing: {selected_file['name']} - Row {current} of {total}")
    
    def on_update_status(self, status_text):
        self.status_label.setText(status_text)
    
    def on_processing_finished(self):
        self.running = False
        self.completed = True
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.status_label.setText("Completed")
        
        # Get selected file and notify parent
        selected_file = self.get_selected_file()
        if selected_file:
            self.parent.update_status(f"Completed: {selected_file['name']}")
        
        # After completing this file, start the next one immediately
        self.parent.process_next_row()
    
    def on_processing_error(self, error_message):
        self.running = False
        self.completed = False
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.status_label.setText(f"Error: {error_message[:40]}...")
        
        # Show error message box
        QMessageBox.warning(self, "Processing Error", error_message)
    
    def lock_controls(self, locked=True):
        """Lock or unlock file selection controls"""
        self.file_dropdown.setEnabled(not locked)
        self.prefix_input.setEnabled(not locked)
        self.refresh_btn.setEnabled(not locked)
    
    def load_files(self):
        """Load files from Google Drive's Web folder, excluding files already selected in other rows"""
        # Disconnect signal temporarily to prevent recursive calls
        try:
            self.file_dropdown.currentIndexChanged.disconnect(self.file_selected)
        except:
            pass  # Signal wasn't connected
            
        # Save current selection
        current_selection = self.file_dropdown.currentText()
        
        # Clear dropdown and add empty item first
        self.file_dropdown.clear()
        self.file_dropdown.addItem("")
        self.status_label.setText("Loading files...")
        
        try:
            # Get files from Google Drive Web folder
            all_files = self.parent.get_drive_web_files()
            
            # Get files that are already selected in other rows
            selected_files = self.parent.get_selected_files()
            
            # Filter out already selected files (except current selection)
            available_files = []
            for file_name in all_files:
                if file_name not in selected_files or file_name == current_selection:
                    available_files.append(file_name)
            
            # Add available files to dropdown
            for file_name in available_files:
                self.file_dropdown.addItem(file_name)
            
            if len(available_files) == 0:
                self.status_label.setText("No files available")
            else:
                self.status_label.setText(f"Found {len(available_files)} files")
            
            # Restore previous selection if available
            if current_selection:
                index = self.file_dropdown.findText(current_selection)
                if index >= 0:
                    self.file_dropdown.setCurrentIndex(index)
                
        except Exception as e:
            print(f"Error loading files: {e}")
            self.status_label.setText("Error loading files")
        
        # Reconnect signal
        self.file_dropdown.currentIndexChanged.connect(self.file_selected)
    
    def file_selected(self):
        """Handle file selection change event"""
        file_name = self.file_dropdown.currentText()
        
        # Save current selection for parent tracking
        if file_name:
            self.selected_file = file_name
            # Try to extract prefix from filename
            self.extract_prefix_from_filename(file_name)
        else:
            self.selected_file = None
        
        # Update other rows' available files - with safety check
        try:
            for i in range(self.parent.scroll_layout.count()):
                item = self.parent.scroll_layout.itemAt(i)
                if item:
                    row = item.widget()
                    if row and row != self and not row.running:
                        row.load_files()
        except Exception as e:
            print(f"Error updating other rows: {e}")
            # Don't crash the application if update fails
    
    def extract_prefix_from_filename(self, filename):
        """Extract prefix from filename (e.g., Globe-605 -> 605)"""
        # Look for pattern like "Globe-605" and extract "605"
        match = re.search(r'[\w]+-(\d+)', filename)
        if match:
            prefix = match.group(1)
            self.prefix_input.setText(prefix)
    
    def get_selected_file(self):
        text = self.file_dropdown.currentText()
        if not text:
            return None
        
        # Handle files from local Web folder
        path = os.path.expanduser(f"~/GoogleDriveMount/Web/{text}")
        return {"name": text, "path": path, "type": "local_file"}
    
    def start_processing(self):
        if self.running:
            return
        
        # Validate inputs
        if not self.file_dropdown.currentText():
            QMessageBox.warning(self, "Error", "Please select a file first")
            return
        
        prefix = self.prefix_input.text().strip()
        if not prefix:
            QMessageBox.warning(self, "Error", "Please enter a Katom prefix")
            return
        
        # Update UI
        self.running = True
        self.completed = False
        self.start_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.progress_bar.setValue(0)
        self.status_label.setText("Starting...")
        
        # Start processing thread
        self.worker_thread = threading.Thread(target=self.process_file)
        self.worker_thread.daemon = True
        self.worker_thread.start()
    
    def stop_processing(self):
        if not self.running:
            return
        
        self.running = False
        self.completed = True
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.status_label.setText("Stopped")
    
    def reset_state(self):
        """Reset the row state completely for reuse"""
        self.running = False
        self.completed = False
        self.progress_bar.setValue(0)
        self.status_label.setText("Ready")
        self.start_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
    
    def process_weight_value(self, value):
        """Process weight values: round up to whole number and add 5"""
        try:
            # Try to extract a number from the string
            # This handles cases like "22.93" or "22.93 lbs"
            number_match = re.search(r'(\d+(\.\d+)?)', str(value))
            if number_match:
                # Extract the number
                number = float(number_match.group(1))
                
                # Round up to nearest whole number
                rounded = math.ceil(number)
                
                # Add 5
                final = rounded + 5
                
                # If the original had units, keep them
                units_match = re.search(r'[^\d.]+$', str(value))
                units = units_match.group(0).strip() if units_match else ""
                
                return f"{final}{' ' + units if units else ''}"
            return value
        except:
            # If any error occurs, return the original value
            return value
    
    def extract_table_data(self, driver):
        """
        Extract table data both as a dictionary of key-value pairs AND as an HTML table.
        Returns a tuple: (specs_dict, specs_html)
        """
        specs_dict = {}
        specs_html = ""
        
        try:
            # Try multiple approaches to find the table
            
            # First, try to get the original table HTML
            specs_tables = driver.find_elements(By.CSS_SELECTOR, "table.table.table-condensed.specs-table")
            
            if not specs_tables:
                # Try generic tables
                specs_tables = driver.find_elements(By.TAG_NAME, "table")
            
            if specs_tables:
                # Extract key-value pairs from the table
                table = specs_tables[0]
                rows = table.find_elements(By.TAG_NAME, "tr")
                
                # Build a clean table with slim styling
                specs_html = '<table class="specs-table" cellspacing="0" cellpadding="4" border="1" style="margin-top:10px;border-collapse:collapse;width:auto;" align="left"><tbody>'
                
                for row in rows:
                    cells = row.find_elements(By.TAG_NAME, "td")
                    if len(cells) >= 2:
                        key = cells[0].text.strip()
                        value = cells[1].text.strip()
                        
                        # Check if this is a weight field and process accordingly
                        if "weight" in key.lower():
                            value = self.process_weight_value(value)
                        
                        # Add to the dictionary
                        if key and not key.lower() in specs_dict:
                            specs_dict[key.lower()] = value
                        
                        # Add to the HTML table
                        specs_html += f'<tr><td style="padding:3px 8px;"><b>{key}</b></td><td style="padding:3px 8px;">{value}</td></tr>'
                
                specs_html += "</tbody></table>"
            
            # If no table found or no HTML extracted, create an HTML table from the data we find
            if not specs_html or specs_html == "":
                # Start building an HTML table
                other_specs = []
                
                # Try to find spec elements in various ways
                
                # Method 1: Look for dedicated spec elements
                spec_rows = driver.find_elements(By.CSS_SELECTOR, ".specs-row, [class*='spec']")
                if spec_rows:
                    for row in spec_rows:
                        key_elem = row.find_elements(By.CSS_SELECTOR, ".spec-key, .spec-name, [class*='key'], [class*='name']")
                        val_elem = row.find_elements(By.CSS_SELECTOR, ".spec-value, .spec-val, [class*='value'], [class*='val']")
                        
                        if key_elem and val_elem:
                            key = key_elem[0].text.strip()
                            value = val_elem[0].text.strip()
                            
                            # Check if this is a weight field and process accordingly
                            if "weight" in key.lower():
                                value = self.process_weight_value(value)
                                
                            if key:
                                other_specs.append((key, value))
                                if not key.lower() in specs_dict:
                                    specs_dict[key.lower()] = value
                
                # Method 2: Look for definition lists
                if not other_specs:
                    dl_elements = driver.find_elements(By.TAG_NAME, "dl")
                    for dl in dl_elements:
                        terms = dl.find_elements(By.TAG_NAME, "dt")
                        definitions = dl.find_elements(By.TAG_NAME, "dd")
                        
                        for i in range(min(len(terms), len(definitions))):
                            key = terms[i].text.strip()
                            value = definitions[i].text.strip()
                            
                            # Check if this is a weight field and process accordingly
                            if "weight" in key.lower():
                                value = self.process_weight_value(value)
                                
                            if key:
                                other_specs.append((key, value))
                                if not key.lower() in specs_dict:
                                    specs_dict[key.lower()] = value
                
                # Method 3: Look for text patterns in all content
                if not other_specs:
                    # Get all elements that might contain specs
                    elements = driver.find_elements(By.CSS_SELECTOR, "p, div, li, span")
                    
                    # Common spec terms to look for - expand this list as needed
                    common_specs = [
                        "manufacturer", "food type", "frypot style", "heat", "hertz", "nema", 
                        "number of", "oil capacity", "phase", "product", "type", "rating", 
                        "special features", "voltage", "warranty", "weight", "dimensions"
                    ]
                    
                    for element in elements:
                        text = element.text.strip()
                        if not text or len(text) > 100:  # Skip empty or very long text
                            continue
                        
                        # Look for patterns like "Key: Value" or "Key - Value"
                        for pattern in [r'([^:]+):\s*(.+)', r'([^-]+)-\s*(.+)']: 
                            match = re.match(pattern, text)
                            if match:
                                key = match.group(1).strip()
                                value = match.group(2).strip()
                                
                                # Check if this is a weight field and process accordingly
                                if "weight" in key.lower():
                                    value = self.process_weight_value(value)
                                
                                # Check if this key is one of our common specs
                                if any(spec in key.lower() for spec in common_specs):
                                    other_specs.append((key, value))
                                    if not key.lower() in specs_dict:
                                        specs_dict[key.lower()] = value
                                    break
                
                # Create HTML table from the data we collected
                if other_specs:
                    specs_html = '<table class="specs-table" cellspacing="0" cellpadding="4" border="1" style="margin-top:10px;border-collapse:collapse;width:auto;" align="left"><tbody>'
                    for key, value in other_specs:
                        specs_html += f'<tr><td style="padding:3px 8px;"><b>{key}</b></td><td style="padding:3px 8px;">{value}</td></tr>'
                    specs_html += "</tbody></table>"
        
        except Exception as e:
            print(f"Error extracting table data: {e}")
        
        return specs_dict, specs_html
    
    def scrape_katom(self, model_number, prefix):
        """Scrape product info from Katom website"""
        # Clean model number
        model_number = ''.join(e for e in model_number if e.isalnum()).upper()
        if model_number.endswith("HC"):
            model_number = model_number[:-2]
        
        url = f"https://www.katom.com/{prefix}-{model_number}.html"
        
        # Set up Selenium
        options = Options()
        options.add_argument('--headless')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument(f'user-agent={UserAgent().random}')
        
        driver = None
        title, description = "Title not found", "Description not found"
        specs_data = {}  # Dictionary to hold spec data
        specs_html = ""  # HTML table for other specs
        video_links = ""  # String to store video links
        item_found = False
        
        try:
            driver = webdriver.Chrome(options=options)
            driver.set_page_load_timeout(30)  # Set timeout to prevent hanging
            driver.get(url)
            
            # Check for 404
            if "404" in driver.title or "not found" in driver.title.lower():
                if driver:
                    driver.quit()
                return title, description, specs_data, specs_html, video_links
            
            # Get title
            try:
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "h1.product-name.mb-0"))
                )
                
                title_element = driver.find_element(By.CSS_SELECTOR, "h1.product-name.mb-0")
                title = title_element.text.strip()
                if title:
                    item_found = True
            except TimeoutException:
                print(f"Timeout waiting for title element on {url}")
                pass
            except Exception as e:
                print(f"Error getting title: {e}")
                pass
            
            # Get description
            if item_found:
                try:
                    tab_content = driver.find_element(By.CLASS_NAME, "tab-content")
                    paragraphs = tab_content.find_elements(By.TAG_NAME, "p")
                    filtered = [
                        f"<p>{p.text.strip()}</p>" for p in paragraphs
                        if p.text.strip() and not p.text.lower().startswith("*free") and "video" not in p.text.lower()
                    ]
                    description = "".join(filtered) if filtered else "Description not found"
                except NoSuchElementException:
                    print(f"Tab content not found on {url}")
                    pass
                except Exception as e:
                    print(f"Error getting description: {e}")
                    pass
                
                # Extract table data as a dictionary AND get HTML table for other specs
                specs_data, specs_html = self.extract_table_data(driver)
                
                # Extract video links - looking specifically for .mp4 sources
                try:
                    # Find source tags with .mp4 files
                    sources = driver.find_elements(By.CSS_SELECTOR, "source[src*='.mp4'], source[type*='video']")
                    for source in sources:
                        src = source.get_attribute("src")
                        if src and src not in video_links:
                            video_links += f"{src}\n"
                    
                    # If no video sources found, look for video elements
                    if not video_links:
                        videos = driver.find_elements(By.TAG_NAME, "video")
                        for video in videos:
                            # Try to get source elements within video tag
                            inner_sources = video.find_elements(By.TAG_NAME, "source")
                            for source in inner_sources:
                                src = source.get_attribute("src")
                                if src and src not in video_links:
                                    video_links += f"{src}\n"
                                    
                    # Last resort - extract video URLs from the page source
                    if not video_links:
                        page_source = driver.page_source
                        # Look for .mp4 URLs in the source
                        mp4_pattern = r'https?://[^"\']+\.mp4'
                        matches = re.findall(mp4_pattern, page_source)
                        for match in matches:
                            if match not in video_links:
                                video_links += f"{match}\n"
                except Exception as e:
                    print(f"Error extracting video links: {e}")
                    
        except Exception as e:
            print(f"Error in scrape_katom: {e}")
            print(traceback.format_exc())
        finally:
            # Ensure driver is always closed, even if an exception occurs
            if driver:
                try:
                    driver.quit()
                except:
                    pass
        
        return title, description, specs_data, specs_html, video_links

    def process_file(self):
        """Process the file in a separate thread"""
        try:
            # Get selected file and prefix
            file_info = self.get_selected_file()
            if not file_info:
                self.signals.error.emit("No file selected")
                return
            
            prefix = self.prefix_input.text().strip()
            
            # Plugin hook: before_process_file
            if hasattr(self.parent, 'plugin_manager'):
                hook_results = self.parent.plugin_manager.execute_hook('before_process_file', self, file_info)
                # Check if any plugin returned False to abort processing
                if hook_results and any(result is False for result in hook_results.values()):
                    self.signals.error.emit("Processing aborted by plugin")
                    return
            
            # Load data from file
            df = self.load_file_data(file_info)
            
            # Find the 'Mfr Model' column
            model_col = None
            for col in df.columns:
                if isinstance(col, str) and col.strip().lower() == 'mfr model':
                    model_col = col
                    break
            
            if not model_col:
                self.signals.error.emit("Missing 'Mfr Model' column in file")
                return
            
            # Define common specification fields we want to extract as individual columns
            # Get from config if available
            if hasattr(self.parent, 'config_manager'):
                common_spec_fields = self.parent.config_manager.get("common_spec_fields")
            else:
                common_spec_fields = [
                    "manufacturer", "food type", "frypot style", "heat", "hertz", "nema", 
                    "number of fry pots", "oil capacity/fryer (lb)", "phase", "product", 
                    "product type", "rating", "special features", "type", "voltage", 
                    "warranty", "weight"
                ]
            
            # Set up output with base columns plus spec fields
            columns = ["Mfr Model", "Title", "Description"]
            for field in common_spec_fields:
                columns.append(field.title())  # Title case the field names for Excel
            
            # Add multiple video link columns (up to 5)
            for i in range(1, 6):  # Video Link 1, Video Link 2, etc.
                columns.append(f"Video Link {i}")
            
            # Initialize output DataFrame with all columns
            self.output_df = pd.DataFrame(columns=columns)
            
            # Set output path from config if available
            if hasattr(self.parent, 'config_manager'):
                output_dir = os.path.expanduser(self.parent.config_manager.get("output", "output_dir"))
                prefix = self.parent.config_manager.get("output", "prefix")
                
                # Create the output path
                filename = f"{prefix}{file_info['name']}.xlsx"
                self.output_path = os.path.join(output_dir, filename)
            else:
                # Use the original output path
                self.output_path = os.path.expanduser(f"~/GoogleDriveMount/Web/Completed/Final/final_{file_info['name']}.xlsx")
            
            # Ensure the output directory exists
            os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
            
            # Save initial file
            self.save_results()
            
            # Process rows
            total_rows = len(df)
            if total_rows == 0:
                self.signals.error.emit("File contains no data rows")
                return
            
            # Update progress at start
            self.signals.update_progress.emit(0, total_rows)
            
            for i, row_data in df.iterrows():
                if not self.running:
                    break
                
                current_row = i + 1
                
                # Get model number
                model = str(row_data[model_col])
                if not model or pd.isna(model):
                    continue
                
                # Process the row
                try:
                    self.signals.update_status.emit(f"Processing model: {model}")
                    
                    title, desc, specs_dict, specs_html, video_links = self.scrape_katom(model, prefix)
                    
                    if title != "Title not found" and "not found" not in title.lower():
                        # Format the description with table at the bottom
                        combined_description = f'<div style="text-align: justify;">{desc}</div>'
                        
                        # Add the specs table below the description if it exists
                        if specs_html and len(specs_html) > 0:
                            combined_description += f'<h3 style="margin-top: 15px;">Specifications</h3>{specs_html}'
                        
                        # Create a row with base data
                        row_data = {
                            "Mfr Model": model,
                            "Title": title,
                            "Description": combined_description
                        }
                        
                        # Add each spec field to its own column if present
                        for field in common_spec_fields:
                            row_data[field.title()] = ""  # Initialize with empty string
                        
                        # Process spec dictionary and add to appropriate columns
                        for key, value in specs_dict.items():
                            # Process weight fields
                            if "weight" in key.lower():
                                value = self.process_weight_value(value)
                                
                            # Try to match spec key to our common fields
                            for field in common_spec_fields:
                                if key.lower() == field.lower() or key.lower() in field.lower():
                                    row_data[field.title()] = value
                                    break
                        
                        # Process video links into separate columns
                        video_list = [link.strip() for link in video_links.strip().split('\n') if link.strip()]
                        for i, link in enumerate(video_list[:5], 1):  # Limit to 5 video links
                            row_data[f"Video Link {i}"] = link
                        
                        # Initialize empty cells for unused video links
                        for i in range(len(video_list) + 1, 6):
                            if i <= 5:  # Ensure we don't exceed our column limit
                                row_data[f"Video Link {i}"] = ""
                        
                        # Add to results DataFrame
                        new_row = pd.DataFrame([row_data])
                        self.output_df = pd.concat([self.output_df, new_row], ignore_index=True)
                        
                        # Save after each successful row
                        self.save_results()
                except Exception as e:
                    print(f"Error processing row {current_row}: {e}")
                    print(traceback.format_exc())
                
                # Update progress
                self.signals.update_progress.emit(current_row, total_rows)
                
                # Small delay to prevent overloading
                time.sleep(0.5)
            
            # Finish up
            if self.running:
                # Plugin hook: after_process_file
                if hasattr(self.parent, 'plugin_manager') and self.output_df is not None:
                    self.parent.plugin_manager.execute_hook('after_process_file', self, self.output_df, self.output_path)
                    
                self.signals.finished.emit()
    
        except Exception as e:
            # Handle unexpected errors
            error_message = str(e)
            print(f"Error in processing: {error_message}")
            print(traceback.format_exc())
            self.signals.error.emit(error_message)
            
    def load_file_data(self, file_info):
        """Load data from a file (Google Sheet or local file)"""
        if file_info['type'] == 'google_sheet':
            # Get sheet by name
            try:
                sheet = self.parent.gc.open(file_info['name'])
                worksheet = sheet.sheet1
                records = worksheet.get_all_records()
                return pd.DataFrame(records)
            except Exception as e:
                print(f"Error loading Google Sheet: {e}")
                raise AppError(f"Failed to load Google Sheet: {str(e)}")
        else:
            # Local file
            path = file_info['path']
            if path.lower().endswith('.csv'):
                return pd.read_csv(path)
            elif path.lower().endswith(('.xlsx', '.xls')):
                return pd.read_excel(path)
            else:
                raise AppError(f"Unsupported file type: {path}")
    
    def save_results(self):
        """Save the current results to the output file with consistent cell heights"""
        if self.output_df is not None and self.output_path:
            try:
                # Ensure the output directory exists
                os.makedirs(os.path.dirname(self.output_path), exist_ok=True)
                
                # First save using pandas to Excel
                self.output_df.to_excel(self.output_path, index=False)
                
                # Then open with openpyxl to adjust cell heights
                workbook = openpyxl.load_workbook(self.output_path)
                worksheet = workbook.active
                
                # Set default row height for all rows
                # Standard default height in Excel is 15 points
                for row in worksheet.iter_rows():
                    worksheet.row_dimensions[row[0].row].height = 15
                
                # Adjust the wrap text settings for the Description column
                for row in worksheet.iter_rows():
                    for cell in row:
                        col_name = worksheet.cell(row=1, column=cell.column).value
                        if col_name == "Description":
                            cell.alignment = Alignment(wrap_text=True)
                
                # Save the modified workbook
                workbook.save(self.output_path)
                workbook.close()  # Explicitly close workbook to release file lock
                
            except Exception as e:
                print(f"Error saving results: {e}")
                print(traceback.format_exc())


class MainWindow(QWidget):
    def __init__(self):
        super().__init__()
        
        # Initialize configuration system
        self.config_manager = ConfigManager()
        
        # Initialize the Google Sheets client
        try:
            self.authenticate_google_drive()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to authenticate with Google Drive: {str(e)}")
            raise
        
        # Set up the UI
        self.setup_ui()
        
        # Add initial row
        self.add_row()
        
        # For sequential processing
        self.processing_queue = []
        self.current_processing_index = -1
        
        # Initialize plugin system
        self.plugin_manager = PluginManager(self)
        self.plugin_manager.discover_plugins()
    
    def authenticate_google_drive(self):
        """Connect to Google Drive API"""
        try:
            creds_path = os.path.expanduser("~/GoogleDriveMount/Web/zapier-454818-4e4abf368f57.json")
            scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
            creds = ServiceAccountCredentials.from_json_keyfile_name(creds_path, scope)
            self.gc = gspread.authorize(creds)
        except Exception as e:
            raise AppError(f"Google Drive authentication failed: {str(e)}")
    
    def get_drive_web_files(self):
        """Get list of available files from local Web folder only"""
        try:
            # Only get files from the local Web folder
            web_folder = os.path.expanduser("~/GoogleDriveMount/Web/")
            local_files = []
            
            if os.path.exists(web_folder):
                print(f"Looking for files in: {web_folder}")
                for filename in os.listdir(web_folder):
                    full_path = os.path.join(web_folder, filename)
                    # Only include files, not directories
                    if os.path.isfile(full_path):
                        if filename.endswith(('.csv', '.xlsx', '.xls')) and not filename.startswith('final_'):
                            local_files.append(filename)
                            print(f"Found file: {filename}")
            
            print(f"Found {len(local_files)} files in local Web folder")
            
            if not local_files:
                # Just for debugging, list all directories and files in parent directory
                parent_dir = os.path.dirname(web_folder)
                if os.path.exists(parent_dir):
                    print(f"Contents of parent directory ({parent_dir}):")
                    for item in os.listdir(parent_dir):
                        print(f"  - {item}")
            
            return sorted(local_files)
        except Exception as e:
            print(f"Error listing files: {e}")
            print(traceback.format_exc())
            return []
    
    def get_selected_files(self):
        """Get a list of files currently selected in any row"""
        selected_files = []
        try:
            # Get all rows in the scroll layout
            for i in range(self.scroll_layout.count()):
                # Safely get the widget at index i
                item = self.scroll_layout.itemAt(i)
                if not item:
                    continue
                    
                # Get the row widget
                row = item.widget()
                # Check if this is a valid SheetRow with a selected file
                if row and isinstance(row, SheetRow) and hasattr(row, 'selected_file') and row.selected_file:
                    selected_files.append(row.selected_file)
                    
        except Exception as e:
            print(f"Error getting selected files: {e}")
            print(traceback.format_exc())
            
        return selected_files
    
    def setup_ui(self):
        """Set up the user interface"""
        # Window properties
        window_title = self.config_manager.get("app", "window_title")
        self.setWindowTitle(window_title)
        window_size = self.config_manager.get("app", "window_size")
        self.setGeometry(100, 100, window_size[0], window_size[1])
        
        # Get colors from config
        primary_color = self.config_manager.get("ui", "button_primary_color")
        secondary_color = self.config_manager.get("ui", "button_secondary_color")
        
        self.setStyleSheet(f"""
            QWidget {{
                background-color: #f0f0f0;
                font-family: Arial;
            }}
            QLabel {{
                color: #333333;
            }}
            QLabel#headerLabel {{
                color: #222222;
                font-size: 18px;
                font-weight: bold;
            }}
            QLabel#statusLabel {{
                color: #333333;
                font-size: 13px;
            }}
            QPushButton {{
                border-radius: 2px;
                padding: 5px 10px;
                font-weight: bold;
                min-height: 30px;
            }}
            QPushButton#actionButton {{
                background-color: {primary_color};
                color: white;
                border: none;
            }}
            QPushButton#actionButton:hover {{
                background-color: #3367d6;
            }}
            QPushButton#actionButton:disabled {{
                background-color: #a5c2f5;
            }}
            QPushButton#secondaryButton {{
                background-color: {secondary_color};
                color: #333333;
                border: 1px solid #cccccc;
            }}
            QPushButton#secondaryButton:hover {{
                background-color: #e5e5e5;
            }}
            QPushButton#dangerButton {{
                background-color: #f5f5f5;
                color: #d32f2f;
                border: 1px solid #ffcdd2;
            }}
            QPushButton#dangerButton:hover {{
                background-color: #ffebee;
            }}
            QScrollArea {{
                border: none;
                background-color: transparent;
            }}
            QScrollBar:vertical {{
                border: none;
                background: #f5f5f5;
                width: 10px;
                border-radius: 5px;
            }}
            QScrollBar::handle:vertical {{
                background: #cccccc;
                min-height: 30px;
                border-radius: 5px;
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                border: none;
                background: none;
            }}
        """)
        
        # Main layout
        main_layout = QVBoxLayout(self)
        main_layout.setContentsMargins(20, 20, 20, 20)
        main_layout.setSpacing(15)
        
        # Header
        header = QLabel("MK Processor", self)
        header.setObjectName("headerLabel")
        header.setAlignment(Qt.AlignCenter)
        
        self.status_label = QLabel("Ready", self)
        self.status_label.setObjectName("statusLabel")
        self.status_label.setAlignment(Qt.AlignCenter)
        
        # Control buttons
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        
        self.start_all_btn = QPushButton("Start All", self)
        self.start_all_btn.setObjectName("actionButton")
        self.start_all_btn.clicked.connect(self.start_all)
        
        self.stop_all_btn = QPushButton("Stop All", self)
        self.stop_all_btn.setObjectName("secondaryButton")
        self.stop_all_btn.clicked.connect(self.stop_all)
        self.stop_all_btn.setEnabled(False)
        
        self.add_row_btn = QPushButton("Add File", self)
        self.add_row_btn.setObjectName("secondaryButton")
        self.add_row_btn.clicked.connect(self.add_row)
        
        self.clear_btn = QPushButton("Clear All", self)
        self.clear_btn.setObjectName("dangerButton")
        self.clear_btn.clicked.connect(self.clear_all)
        
        # Add settings button
        self.settings_btn = QPushButton("Settings", self)
        self.settings_btn.setObjectName("secondaryButton")
        self.settings_btn.clicked.connect(self.show_settings)
        
        # Add plugins button
        self.plugins_btn = QPushButton("Manage Plugins", self)
        self.plugins_btn.setObjectName("secondaryButton")
        self.plugins_btn.clicked.connect(self.show_plugin_manager)
        
        button_layout.addWidget(self.start_all_btn)
        button_layout.addWidget(self.stop_all_btn)
        button_layout.addWidget(self.add_row_btn)
        button_layout.addWidget(self.clear_btn)
        button_layout.addWidget(self.settings_btn)
        button_layout.addWidget(self.plugins_btn)
        
        # Scroll area for rows
        self.scroll_area = QScrollArea(self)
        self.scroll_area.setWidgetResizable(True)
        
        self.scroll_content = QWidget()
        self.scroll_layout = QVBoxLayout(self.scroll_content)
        self.scroll_layout.setAlignment(Qt.AlignTop)
        self.scroll_layout.setSpacing(10)
        self.scroll_layout.setContentsMargins(5, 5, 5, 5)
        
        self.scroll_area.setWidget(self.scroll_content)
        
        # Add widgets to main layout
        main_layout.addWidget(header)
        main_layout.addWidget(self.status_label)
        main_layout.addLayout(button_layout)
        main_layout.addWidget(self.scroll_area, 1)  # 1 = stretch factor
    
    def show_settings(self):
        """Show the settings dialog"""
        dialog = SettingsDialog(self.config_manager, self)
        dialog.exec_()
    
    def show_plugin_manager(self):
        """Show the plugin manager dialog"""
        dialog = PluginManagerDialog(self.plugin_manager, self)
        dialog.exec_()
    
    def add_row(self):
        """Add a new row to the interface"""
        try:
            # Create new SheetRow
            row = SheetRow(len(self.scroll_layout), self)
            
            # Apply enhancements
            row = create_webscraper_wrapper(row)
            row = enhance_save_results(row)
            
            # Force UI update before adding it
            QApplication.processEvents()
            
            # Add to layout
            self.scroll_layout.addWidget(row)
            
            # Force another UI update to ensure styling is applied
            QApplication.processEvents()
            
            # Scroll to the bottom to show the new row
            QTimer.singleShot(100, lambda: self.scroll_area.verticalScrollBar().setValue(
                self.scroll_area.verticalScrollBar().maximum()))
                
            # Update all other rows to refresh available files
            QTimer.singleShot(500, self.refresh_all_rows)
        except Exception as e:
            print(f"Error adding row: {e}")
            print(traceback.format_exc())
            QMessageBox.warning(self, "Error", f"Error adding new row: {str(e)}")
    
    def refresh_all_rows(self):
        """Refresh file lists in all rows"""
        try:
            for i in range(self.scroll_layout.count()):
                item = self.scroll_layout.itemAt(i)
                if item:
                    row = item.widget()
                    if row and not row.running:
                        row.load_files()
        except Exception as e:
            print(f"Error refreshing rows: {e}")
            # Don't crash the application if refresh fails
    
    def clear_all(self):
        """Remove all rows"""
        # Confirm if any rows are running
        running_found = False
        for i in range(self.scroll_layout.count()):
            item = self.scroll_layout.itemAt(i)
            if item:
                row = item.widget()
                if row and row.running:
                    running_found = True
                    break
        
        if running_found:
            reply = QMessageBox.question(
                self, "Confirm", 
                "Processing is currently running. Stop and clear all?",
                QMessageBox.Yes | QMessageBox.No, 
                QMessageBox.No
            )
            if reply != QMessageBox.Yes:
                return
        
        # Stop all running processes
        self.stop_all()
        
        # Clear all rows
        while self.scroll_layout.count():
            item = self.scroll_layout.takeAt(0)
            if item and item.widget():
                item.widget().deleteLater()
        
        # Reset status
        self.update_status("Ready")
        
        # Add one empty row
        self.add_row()
    
    def start_all(self):
        """Start processing all rows sequentially without delay between rows"""
        valid_rows = []
        for i in range(self.scroll_layout.count()):
            item = self.scroll_layout.itemAt(i)
            if not item:
                continue
                
            row = item.widget()
            if row and row.file_dropdown.currentText() and row.prefix_input.text().strip():
                # Reset row state to ensure clean start
                if hasattr(row, 'reset_state'):
                    row.reset_state()
                
                # Lock file selection controls
                if hasattr(row, 'lock_controls'):
                    row.lock_controls(True)
                    
                valid_rows.append(row)
        
        if not valid_rows:
            QMessageBox.warning(self, "Error", "Please add at least one file with a prefix")
            return
        
        # Update UI
        self.start_all_btn.setEnabled(False)
        self.stop_all_btn.setEnabled(True)
        self.update_status("Starting sequential processing...")
        
        # Set up queue for sequential processing
        self.processing_queue = valid_rows
        self.current_processing_index = -1
        
        # Start the first row
        QTimer.singleShot(100, self.process_next_row)
    
    def process_next_row(self):
        """Process the next row in the queue immediately"""
        if not self.stop_all_btn.isEnabled():
            return
        
        # Increment index
        self.current_processing_index += 1
        
        # If we reached the end of the queue, we're done
        if self.current_processing_index >= len(self.processing_queue):
            self.update_status("All processing completed")
            self.start_all_btn.setEnabled(True)
            self.stop_all_btn.setEnabled(False)
            return
        
        # Get the next row to process
        row = self.processing_queue[self.current_processing_index]
        
        # Update status
        file_info = row.get_selected_file()
        if file_info:
            self.update_status(f"Starting file {self.current_processing_index + 1} of {len(self.processing_queue)}: {file_info['name']}")
        
        # Start processing immediately
        self._start_row(row)
    
    def _start_row(self, row):
        """Start processing the row"""
        # Only start if not already running or completed
        if not row.running and not row.completed:
            # Now start processing
            row.start_processing()
    
    def stop_all(self):
        """Stop all running processes"""
        # Clear the processing queue
        self.processing_queue = []
        
        # Reset all rows
        for i in range(self.scroll_layout.count()):
            item = self.scroll_layout.itemAt(i)
            if not item:
                continue
                
            row = item.widget()
            if row:
                # Stop processing if running
                if row.running:
                    row.stop_processing()
                row.status_label.setText("Stopped")
                
                # Unlock file selection controls
                if hasattr(row, 'lock_controls'):
                    row.lock_controls(False)
        
        # Update UI state
        self.start_all_btn.setEnabled(True)
        self.stop_all_btn.setEnabled(False)
        self.update_status("Stopped")
    
    def update_status(self, message):
        """Update the status label"""
        self.status_label.setText(message)


def main():
    # Set up application
    app = QApplication(sys.argv)
    
    try:
        # Create and show the main window
        window = MainWindow()
        window.show()
        
        # Print info about loaded plugins
        if hasattr(window, 'plugin_manager'):
            print(f"Loaded plugins: {list(window.plugin_manager.plugins.keys())}")
        
        # Run the application
        sys.exit(app.exec_())
    except Exception as e:
        # Show error dialog for unexpected errors
        print(f"Application error: {e}")
        print(traceback.format_exc())
        
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setWindowTitle("Application Error")
        msg.setText(f"The application failed to start: {str(e)}")
        msg.setDetailedText(traceback.format_exc())
        msg.exec_()
        
        sys.exit(1)


if __name__ == "__main__":
    main()
